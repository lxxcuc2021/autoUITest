#!/usr/bin/env python
# _*_ coding:utf-8 _*_

from openpyxl import load_workbook
import os
from config import globalconfig

data_path = globalconfig.data_path


class ReadExcel(object):
	"""
	读取excel
	"""

	def __init__(self, excel_path=None, sheet_name=None):
		# excel路径为空取默认值
		if excel_path is None:
			self.excel_path = os.path.join(data_path, 'casedata.xlsx')
		else:
			self.excel_path = excel_path
		# 工作表参数为空默认为Sheet1
		if sheet_name is None:
			self.sheet_name = 'Sheet1'
		else:
			self.sheet_name = sheet_name

		self.wb = load_workbook(self.excel_path)
		self.sheet = self.wb[self.sheet_name]

	def excel_reader(self):
		rows = self.sheet.rows
		row_num = self.sheet.max_row
		col_num = self.sheet.max_column

		if row_num <= 1:
			print('总行数小于1，没有数据')
		else:
			data = []
			for row in rows:
				row_data = []
				for i in range(col_num):
					row_data.append(row[i].value)
				data.append(row_data)
			return data


# if __name__ == '__main__':
# 	excel = ExcelUtil(sheet_name='Login')
# 	file_data = excel.excel_reader()
# 	print(file_data)
